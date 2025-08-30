from datetime import date
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Q
from django.utils import timezone
import csv

from django.contrib.auth import get_user_model
from .models import AuditLog, AuditAction

User = get_user_model()

def is_staff(u): return u.is_authenticated and u.is_staff

@login_required
@user_passes_test(is_staff)
def audit_log(request):
    # Filters
    q = (request.GET.get("q") or "").strip()
    action = request.GET.get("action") or ""
    user_id = request.GET.get("user") or ""
    subject_id = request.GET.get("subject") or ""
    shift_id = request.GET.get("shift") or ""
    booking_id = request.GET.get("booking") or ""
    start = request.GET.get("start") or ""
    end = request.GET.get("end") or ""

    qs = AuditLog.objects.select_related("actor", "subject", "shift", "booking")

    if q:
        qs = qs.filter(
            Q(message__icontains=q) |
            Q(actor__username__icontains=q) |
            Q(subject__username__icontains=q)
        )
    if action:
        qs = qs.filter(action=action)
    if user_id.isdigit():
        qs = qs.filter(actor_id=int(user_id))
    if subject_id.isdigit():
        qs = qs.filter(subject_id=int(subject_id))
    if shift_id.isdigit():
        qs = qs.filter(shift_id=int(shift_id))
    if booking_id.isdigit():
        qs = qs.filter(booking_id=int(booking_id))
    if start:
        qs = qs.filter(at__date__gte=start)
    if end:
        qs = qs.filter(at__date__lte=end)

    qs = qs.order_by("-at")[:1000]

    # Export
    fmt = (request.GET.get("format") or "").lower()
    if fmt == "csv":
        return _audit_csv(qs, start, end)
    if fmt == "xlsx":
        try:
            return _audit_xlsx(qs, start, end)
        except ImportError:
            # Fallback message in page in real app; here we just give CSV
            return _audit_csv(qs, start, end)

    context = {
        "logs": qs,
        "actions": AuditAction.choices,
        "users": User.objects.order_by("username")[:300],
        "q": q, "action_filter": action, "user_filter": user_id,
        "subject_filter": subject_id, "shift_filter": shift_id, "booking_filter": booking_id,
        "start": start, "end": end,
    }
    return render(request, "audit/log.html", context)


def _audit_csv(qs, start, end):
    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = f'attachment; filename="audit_{start or "all"}_{end or "all"}.csv"'
    w = csv.writer(resp)
    w.writerow(["Timestamp", "Action", "Actor", "Subject", "Shift", "Booking", "Message", "Extra JSON"])
    for a in qs:
        w.writerow([
            timezone.localtime(a.at).strftime("%Y-%m-%d %H:%M:%S"),
            a.get_action_display(),
            getattr(a.actor, "username", "") if a.actor_id else "",
            getattr(a.subject, "username", "") if a.subject_id else "",
            getattr(a.shift, "title", "") if a.shift_id else "",
            a.booking_id or "",
            (a.message or "").replace("\n", " "),
            a.extra,
        ])
    return resp


def _audit_xlsx(qs, start, end):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit"

    headers = ["Timestamp", "Action", "Actor", "Subject", "Shift", "Booking", "Message", "Extra JSON"]
    ws.append(headers)
    for a in qs:
        ws.append([
            timezone.localtime(a.at).strftime("%Y-%m-%d %H:%M:%S"),
            a.get_action_display(),
            getattr(a.actor, "username", "") if a.actor_id else "",
            getattr(a.subject, "username", "") if a.subject_id else "",
            getattr(a.shift, "title", "") if a.shift_id else "",
            a.booking_id or "",
            a.message or "",
            str(a.extra or {}),
        ])
    # widen
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="audit_{start or "all"}_{end or "all"}.xlsx"'
    return resp
