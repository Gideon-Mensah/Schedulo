# shifts/views.py
from __future__ import annotations

import base64
import csv
import json
import secrets
import string
from datetime import date, timedelta, datetime, time
from io import BytesIO

import requests
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.views import LoginView
from django.db import transaction, IntegrityError, DatabaseError, connection
from django.db.models import Count, F, Q
from django.db.models.functions import TruncDate
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_control, never_cache
from django.views.decorators.http import require_POST
from core.tenant import org_context, get_current_org


from core.org_context import org_context

from .utils import user_is_compliant_for_role

from .forms import AdminComplianceUploadForm, AdminUserCreateForm, ShiftForm, UserAvailabilityForm, HolidayRequestForm, AdminHolidayResponseForm
from .models import ComplianceDocType, Shift, ShiftBooking, UserAvailability, HolidayRequest
from .models import ComplianceDocument
from .utils import log_audit
from .models import AuditAction
import logging, traceback
logger = logging.getLogger(__name__)

# ---------- Helpers ----------
def is_admin(user):
    return user.is_authenticated and user.is_staff

User = get_user_model()

def _active_tenant(request):
    return getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)

# ---------- Admin: create / list shifts ----------


logger = logging.getLogger("django")
@login_required
def create_shift(request):
    logger.debug(f"[DEBUG] request.tenant: {getattr(request, 'tenant', None)}")
    """
    Create a Shift for the current user's organization.
    - Requires the user to have a profile with an organization.
    - Uses atomic transaction so the audit log and shift save are consistent.
    """
    # 1) Preconditions: profile org + tenant present and matching
    try:
        user_org = request.user.profile.organization
    except AttributeError:
        messages.error(request, "You must belong to an organization to create shifts.")
        return redirect("profile_setup")  # or render the form page; adjust to your flow

    tenant = getattr(request, "tenant", None)
    if tenant is None:
        messages.error(request, "No active workspace detected. Please select an organization.")
        return redirect("home")

    if getattr(user_org, "pk", None) != getattr(tenant, "pk", None):
        messages.error(request, "Your profile organization doesn't match the active workspace.")
        return redirect("admin_manage_shifts")

    if request.method == "POST":
        form = ShiftForm(request.POST)
        if form.is_valid():
            shift = form.save(commit=False)
            shift.organization = tenant

            try:
                with transaction.atomic():
                    with org_context(tenant):
                        shift.save()
                        logger.info(
                            "Shift saved id=%s org_id=%s db_alias=%s vendor=%s",
                            shift.id, shift.organization_id, shift
                        )

                    # 2) Defer audit so audit errors NEVER roll back the shift
                    def _after_commit():
                        try:
                            log_audit(
                                actor=request.user,
                                action=AuditAction.SHIFT_CREATED,
                                shift=shift,
                                message=f"Shift '{shift.title}' on {shift.date} created.",
                                role=shift.role,
                                start=str(shift.start_time),
                                end=str(shift.end_time),
                            )
                        except Exception as audit_exc:
                            logger.exception("Audit log failed post-commit: %s", audit_exc)

                    transaction.on_commit(_after_commit)

                messages.success(request, "Shift created successfully.")
                return redirect("admin_manage_shifts")

            except (IntegrityError, DatabaseError) as db_exc:
                logger.exception("DB error creating shift")
                msg = f"Database error while creating the shift: {db_exc}" if settings.DEBUG else "Database error while creating the shift."
                messages.error(request, msg)

            except Exception as exc:
                # 3) Show full traceback in DEBUG so we see the real cause in logs
                tb = traceback.format_exc()
                logger.error("Unexpected error creating shift: %s\n%s", exc, tb)
                msg = f"Unexpected error: {exc}" if settings.DEBUG else "Something went wrong while creating the shift. Please try again."
                messages.error(request, msg)
        else:
            messages.error(request, "Please fix the errors below.")
    else:
        form = ShiftForm()
    org_name = getattr(getattr(request.user, 'profile', None), 'organization', None)
    org_name = getattr(org_name, 'name', "") if org_name else ""
    return render(request, "create_shift.html", {"form": form, "org_name": org_name})

@user_passes_test(is_admin)
def list_shifts(request):
    # ---- resolve active tenant/org ----
    tenant = getattr(request, "tenant", None)
    if tenant is None:
        tenant = getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    # ---- filter for upcoming shifts only ----
    now = timezone.localtime()
    today = now.date()
    current_time = now.time()
    
    # upcoming = strictly in future OR later today (not finished)
    future_q = (
        Q(date__gt=today) |
        (Q(date=today) & (
            Q(end_time__gt=current_time) |
            Q(end_time__isnull=True, start_time__gt=current_time) |
            Q(start_time__isnull=True, end_time__isnull=True)
        ))
    )

    shifts = (
        Shift._base_manager
        .filter(future_q, organization=tenant)
        .annotate(booked_count=Count("bookings"))
        .order_by("date", "start_time")
    )
    
    # Calculate statistics
    total_shifts = shifts.count()
    available_shifts = sum(1 for shift in shifts if shift.booked_count < shift.max_staff)
    full_shifts = sum(1 for shift in shifts if shift.booked_count >= shift.max_staff)
    
    context = {
        'shifts': shifts,
        'total_shifts': total_shifts,
        'available_shifts': available_shifts,
        'full_shifts': full_shifts,
    }
    
    return render(request, "list_shifts.html", context)

# ---------- Staff: Shifts & bookings ----------
@login_required
def available_shifts(request):
    # ---- resolve active tenant/org ----
    tenant = getattr(request, "tenant", None)
    if tenant is None:
        tenant = getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    now = timezone.localtime()
    today = now.date()
    current_time = now.time()

    # upcoming: future dates, or later today if not finished
    future_q = (
        Q(date__gt=today) |
        (
            Q(date=today) & (
                Q(end_time__gt=current_time) |
                Q(end_time__isnull=True, start_time__gt=current_time) |
                Q(start_time__isnull=True, end_time__isnull=True)
            )
        )
    )

    # shifts the user has already booked (org-scoped explicitly)
    booked_shift_ids = (
        ShiftBooking._base_manager
        .filter(user=request.user, organization=tenant)
        .values_list("shift_id", flat=True)
    )

    # available shifts = org-scoped upcoming, not already booked, not full
    shifts = (
        Shift._base_manager
        .filter(future_q, organization=tenant)
        .exclude(id__in=booked_shift_ids)
        .annotate(booked_total=Count("bookings", distinct=True))
        .filter(booked_total__lt=F("max_staff"))
        .order_by("date", "start_time")
    )

    return render(request, "available_shifts.html", {"shifts": shifts})


@login_required
def book_shift(request, shift_id):
    shift = get_object_or_404(Shift, id=shift_id)

    # already full?
    if not shift.has_space:
        messages.warning(request, f"'{shift.title}' is already full.")
        return redirect("available_shifts")

    # already booked?
    if ShiftBooking.objects.filter(user=request.user, shift=shift).exists():
        messages.info(request, f"You have already booked '{shift.title}'.")
        return redirect("available_shifts")

    booking = ShiftBooking.objects.create(user=request.user, shift=shift, organization=shift.organization)
    messages.success(request, f"You have successfully booked '{shift.title}'.")
    
    # For Audit log
    log_audit(actor=request.user, subject=request.user, action=AuditAction.BOOKING_CREATED,
          shift=shift, booking=booking,
          message=f"{request.user.username} booked '{shift.title}'.")
    
    return redirect("available_shifts")


@login_required
def my_bookings(request):
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    now = timezone.localtime()
    today = now.date()
    current_time = now.time()

    # upcoming: strictly future OR later today (not finished)
    future_q = (
        Q(shift__date__gt=today) |
        (
            Q(shift__date=today) & (
                Q(shift__end_time__gt=current_time) |
                Q(shift__end_time__isnull=True, shift__start_time__gt=current_time) |
                Q(shift__start_time__isnull=True, shift__end_time__isnull=True)
            )
        )
    )

    # in-progress or needs action: clocked in but not clocked out
    needs_clock_out_q = Q(clock_in_at__isnull=False, clock_out_at__isnull=True)

    bookings = (
        ShiftBooking._base_manager
        .select_related("shift")
        .filter(user=request.user, organization=tenant)
        # Hide COMPLETED ones
        .exclude(clock_in_at__isnull=False, clock_out_at__isnull=False)
        # Show upcoming OR still needs clock-out
        .filter(future_q | needs_clock_out_q)
        .order_by("shift__date", "shift__start_time", "id")
    )

    # keep your existing UI flags
    for b in bookings:
        sh = b.shift
        sd = datetime.combine(sh.date, sh.start_time or time.min, tzinfo=timezone.get_current_timezone())
        ed = datetime.combine(sh.date, sh.end_time or time.max, tzinfo=timezone.get_current_timezone())
        grace_deadline = sd + timedelta(minutes=30)

        b.show_clock_in = False
        b.show_clock_out = False

        if b.clock_in_at is None:
            # allow clock-in up to 30 min after start while shift not ended
            if sd <= now <= grace_deadline and now < ed:
                b.show_clock_in = True
            elif sd < now <= ed and now > grace_deadline:
                # after grace window, let them clock out directly (late start)
                b.show_clock_out = True
        else:
            if b.clock_out_at is None:
                b.show_clock_out = True

    return render(request, "my_bookings.html", {"bookings": bookings})


@login_required
def completed_shifts(request):
    """User tab: completed but not yet paid."""
    # Use the same tenant resolution as my_bookings
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    bookings = (
        ShiftBooking._base_manager
        .select_related("shift")
        .filter(
            user=request.user,
            organization=tenant,
            clock_in_at__isnull=False,
            clock_out_at__isnull=False,
            paid_at__isnull=True,     # NOT paid yet
        )
        .order_by("-shift__date", "-id")
    )
    return render(request, "shifts/completed_shifts.html", {"bookings": bookings})

@login_required
def past_shifts(request):
    """User tab: completed and paid (historical)."""
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    bookings = (
        ShiftBooking._base_manager
        .select_related("shift")
        .filter(
            user=request.user,
            organization=tenant,
            clock_in_at__isnull=False,
            clock_out_at__isnull=False,
            paid_at__isnull=False,    # already paid
        )
        .order_by("-shift__date", "-id")
    )
    return render(request, "shifts/past_shifts.html", {"bookings": bookings})



@login_required
def cancel_booking(request, booking_id):
    # Use _base_manager to bypass tenant filtering, then manually check tenant
    booking = get_object_or_404(ShiftBooking._base_manager, id=booking_id, user=request.user)
    
    # Verify the booking belongs to the user's organization
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant and booking.organization_id != getattr(tenant, "id", None):
        messages.error(request, "Booking not found or access denied.")
        return redirect("my_bookings")

    if request.method != "POST":
        messages.error(request, "Invalid request.")
        return redirect("my_bookings")

    if booking.shift.is_past:
        messages.warning(
            request,
            f"'{booking.shift.title}' has already started/finished and cannot be cancelled.",
        )
        return redirect("my_bookings")

    booking.delete()
    messages.success(request, f"You have successfully canceled '{booking.shift.title}'.")
    
    # For Audit log
    log_audit(actor=request.user, subject=booking.user, action=AuditAction.BOOKING_CANCELLED,
          shift=booking.shift, booking=None,
          message=f"{request.user.username} cancelled booking for '{booking.shift.title}'.")
    return redirect("my_bookings")

# ---------- Calendar View ----------
@login_required
def my_calendar(request):
    """Calendar view showing user's booked shifts, availability, and holidays"""
    import calendar
    from .models import UserAvailability, HolidayRequest
    
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    # Get month and year from query params or use current
    now = timezone.localtime()
    try:
        year = int(request.GET.get('year', now.year))
        month = int(request.GET.get('month', now.month))
    except (ValueError, TypeError):
        year, month = now.year, now.month

    # Ensure valid month/year ranges
    if month < 1:
        month = 12
        year -= 1
    elif month > 12:
        month = 1
        year += 1

    # Calculate previous and next month for navigation
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    # Get first and last day of the month
    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)

    # Get all bookings for the month
    bookings = (
        ShiftBooking._base_manager
        .select_related("shift")
        .filter(
            user=request.user,
            organization=tenant,
            shift__date__gte=first_day,
            shift__date__lte=last_day
        )
        .order_by("shift__date", "shift__start_time")
    )

    # Get availability for the month
    availability = (
        UserAvailability._base_manager
        .filter(
            user=request.user,
            organization=tenant,
            date__gte=first_day,
            date__lte=last_day
        )
        .order_by("date", "start_time")
    )

    # Get holiday requests for the month
    holidays = (
        HolidayRequest._base_manager
        .filter(
            user=request.user,
            organization=tenant,
            start_date__lte=last_day,
            end_date__gte=first_day
        )
        .order_by("start_date")
    )

    # Group data by date
    bookings_by_date = {}
    for booking in bookings:
        shift_date = booking.shift.date
        if shift_date not in bookings_by_date:
            bookings_by_date[shift_date] = []
        bookings_by_date[shift_date].append(booking)

    availability_by_date = {}
    for avail in availability:
        if avail.date not in availability_by_date:
            availability_by_date[avail.date] = []
        availability_by_date[avail.date].append(avail)

    # Create holiday date set for easy lookup
    holiday_dates = set()
    holiday_info_by_date = {}
    for holiday in holidays:
        current_date = holiday.start_date
        while current_date <= holiday.end_date:
            if first_day <= current_date <= last_day:
                holiday_dates.add(current_date)
                holiday_info_by_date[current_date] = holiday
            current_date += timedelta(days=1)

    # Generate calendar
    cal = calendar.Calendar(firstweekday=0)  # Monday = 0
    month_days = cal.monthdayscalendar(year, month)
    
    # Create calendar data structure
    calendar_weeks = []
    for week in month_days:
        week_data = []
        for day in week:
            if day == 0:
                week_data.append({'day': None, 'bookings': [], 'availability': [], 'holiday': None})
            else:
                day_date = date(year, month, day)
                day_bookings = bookings_by_date.get(day_date, [])
                day_availability = availability_by_date.get(day_date, [])
                day_holiday = holiday_info_by_date.get(day_date)
                
                week_data.append({
                    'day': day,
                    'date': day_date,
                    'bookings': day_bookings,
                    'availability': day_availability,
                    'holiday': day_holiday,
                    'is_today': day_date == now.date(),
                    'is_past': day_date < now.date(),
                    'is_holiday': day_date in holiday_dates
                })
        calendar_weeks.append(week_data)

    context = {
        'year': year,
        'month': month,
        'month_name': calendar.month_name[month],
        'calendar_weeks': calendar_weeks,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
        'today': now.date(),
        'weekdays': ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
        'bookings': bookings,
        'availability': availability,
        'holidays': holidays
    }
    
    return render(request, "shifts/calendar.html", context)

# ---------- Login (no-cache) ----------
@method_decorator(
    [never_cache, cache_control(no_cache=True, no_store=True, must_revalidate=True)],
    name="dispatch",
)
class NoCacheLoginView(LoginView):
    template_name = "login.html"
    redirect_authenticated_user = True

    def get_success_url(self):
        return reverse("account_profile")

    def render_to_response(self, context, **response_kwargs):
        resp = super().render_to_response(context, **response_kwargs)
        resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp["Pragma"] = "no-cache"
        resp["Expires"] = "0"
        return resp


@login_required
@cache_control(no_store=True, no_cache=True, must_revalidate=True)
@never_cache
def home(request):
    # send staff to dashboard, others to the booking page
    if request.user.is_staff:
        return redirect("admin_dashboard")
    return redirect("available_shifts")

# ---------- Location & postcode helpers ----------
def _parse_coords_from_json(request):
    """Expecting JSON body: { 'lat': <float>, 'lng': <float> }"""
    try:
        data = json.loads(request.body.decode("utf-8"))
        return float(data["lat"]), float(data["lng"])
    except Exception:
        return None

def _resolve_postcode(lat: float, lng: float) -> str | None:
    """Reverse-geocode with OpenStreetMap Nominatim to get a postcode."""
    try:
        resp = requests.get(
            "https://nominatim.openstreetmap.org/reverse",
            params={"format": "jsonv2", "lat": lat, "lon": lng, "zoom": 18, "addressdetails": 1},
            headers={"User-Agent": "ScheduloApp/1.0 (contact: admin@example.com)"},
            timeout=6,
        )
        if resp.status_code != 200:
            return None
        js = resp.json()
        addr = js.get("address") or {}
        return addr.get("postcode") or addr.get("postal_code") or addr.get("ISO3166-2-lvl4")
    except Exception:
        return None

def _clock_json(ok: bool, msg: str, status: int = 200):
    return JsonResponse({"ok": ok, "msg": msg}, status=status)

def _parse_json(request):
    try:
        return json.loads(request.body.decode("utf-8"))
    except Exception:
        return {}

def _save_signature_from_dataurl(data_url: str, filename_prefix: str = "signature"):
    """
    Accepts 'data:image/png;base64,AAAA...' and returns a ContentFile.
    """
    if not data_url or not data_url.startswith("data:image"):
        return None
    try:
        header, b64 = data_url.split(",", 1)
        ext = "png"
        if "image/jpeg" in header:
            ext = "jpg"
        data = base64.b64decode(b64)
        from django.core.files.base import ContentFile

        return ContentFile(data, name=f"{filename_prefix}.{ext}")
    except Exception:
        return None

# ---------- Clock in / out ----------
@require_POST
@login_required
def clock_in(request, booking_id):
    try:
        # Use _base_manager to bypass tenant filtering, then manually check tenant
        booking = get_object_or_404(ShiftBooking._base_manager, id=booking_id, user=request.user)
        
        # Verify the booking belongs to the user's organization
        tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
        if tenant and booking.organization_id != getattr(tenant, "id", None):
            return _clock_json(False, "Booking not found or access denied.", status=404)
        
        now = timezone.localtime()

        # If your existing guard rejects, allow within 30-minute grace after start
        if not booking.can_clock_in(now=now):
            sh = booking.shift
            sd = datetime.combine(sh.date, sh.start_time or time.min, tzinfo=timezone.get_current_timezone())
            ed = datetime.combine(sh.date, sh.end_time or time.max, tzinfo=timezone.get_current_timezone())
            from datetime import timedelta
            if not (sd <= now <= min(sd + timedelta(minutes=30), ed)):
                return _clock_json(False, "Clock-in not allowed at this time.", status=400)

        coords = _parse_coords_from_json(request)
        if not coords:
            return _clock_json(False, "Missing or invalid coordinates.", status=400)
        lat, lng = coords

        resolved_pc = _resolve_postcode(lat, lng)  # may be None
        target_pc_raw = booking.shift.allowed_postcode or ""

        if target_pc_raw:
            if not resolved_pc:
                return _clock_json(False, "Could not determine your postcode from location. Please enable precise location (GPS) and try again.", status=422)

            def _norm(s): return "".join(ch for ch in (s or "").upper() if ch.isalnum())
            def _outward(s):
                n = _norm(s); return n[:-3] if len(n) > 3 else n

            if not (_norm(resolved_pc) == _norm(target_pc_raw) or _outward(resolved_pc) == _outward(target_pc_raw)):
                return _clock_json(False, f"Postcode mismatch. Detected: {resolved_pc}; expected: {target_pc_raw}.", status=403)

        # ✅ actually persist the clock-in
        if not booking.clock_in_at:
            booking.clock_in_at = timezone.now()
            booking.clock_in_lat = lat
            booking.clock_in_lng = lng
            booking.clock_in_postcode = resolved_pc or None
            booking.save(update_fields=["clock_in_at", "clock_in_lat", "clock_in_lng", "clock_in_postcode"])

        log_audit(actor=request.user, subject=request.user, action=AuditAction.CLOCK_IN,
                  shift=booking.shift, booking=booking,
                  message="Clock in recorded.",
                  detected_postcode=resolved_pc, lat=lat, lng=lng)

        messages.success(request, "Clock-in recorded.")
        return _clock_json(True, "Clock-in successful.")
        
    except Exception as e:
        # Log the error for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Clock-in error for booking {booking_id}: {str(e)}", exc_info=True)
        return _clock_json(False, "An error occurred during clock-in. Please try again.", status=500)

@require_POST
@login_required
def clock_out(request, booking_id):
    try:
        # Use _base_manager to bypass tenant filtering, then manually check tenant
        booking = get_object_or_404(ShiftBooking._base_manager, id=booking_id, user=request.user)
        
        # Verify the booking belongs to the user's organization
        tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
        if tenant and booking.organization_id != getattr(tenant, "id", None):
            return _clock_json(False, "Booking not found or access denied.", status=404)
        
        now = timezone.localtime()

        if not booking.can_clock_out(now=now):
            # Still allow if we're inside the shift window
            sh = booking.shift
            sd = datetime.combine(sh.date, sh.start_time or time.min, tzinfo=timezone.get_current_timezone())
            ed = datetime.combine(sh.date, sh.end_time or time.max, tzinfo=timezone.get_current_timezone())
            if not (sd <= now <= ed):
                return _clock_json(False, "Clock-out not allowed right now.", status=400)

            # If they never clocked in and we're past the 30-min grace, backfill a sensible "in" time
        if booking.clock_in_at is None:
            from datetime import timedelta
            sd = datetime.combine(booking.shift.date, booking.shift.start_time or time.min, tzinfo=timezone.get_current_timezone())
            grace_deadline = sd + timedelta(minutes=30)
            # Use the grace deadline, capped at 'now' just in case
            booking.clock_in_at = min(grace_deadline, now)
            # Optional: annotate that this was an auto backfill
            booking.clock_out_note = (booking.clock_out_note or "")
            booking.clock_out_note += (("\n" if booking.clock_out_note else "") + "[System] Auto-set clock-in at start+30m due to late clock-in.")

        payload = _parse_json(request)

        coords = _parse_coords_from_json(request)
        if not coords:
            return _clock_json(False, "Missing or invalid coordinates.", status=400)
        lat, lng = coords

        resolved_pc = _resolve_postcode(lat, lng)
        target_pc_raw = booking.shift.allowed_postcode or ""

        if target_pc_raw:
            if not resolved_pc:
                return _clock_json(
                    False,
                    "Could not determine your postcode from location. Please enable precise location (GPS) and try again.",
                    status=422,
                )

            def _norm(s): return "".join(ch for ch in (s or "").upper() if ch.isalnum())
            def _outward(s):
                n = _norm(s)
                return n[:-3] if len(n) > 3 else n

            if not (_norm(resolved_pc) == _norm(target_pc_raw) or _outward(resolved_pc) == _outward(target_pc_raw)):
                return _clock_json(
                    False, f"Postcode mismatch. Detected: {resolved_pc}; expected: {target_pc_raw}.", status=403
                )

        if booking.clock_out_at:
            return _clock_json(True, "Already clocked out.")

        # Optional extras
        note = (payload.get("note") or "").strip()
        supervisor_name = (payload.get("supervisor_name") or "").strip()
        sig_data_url = payload.get("signature_data_url") or ""

        booking.clock_out_at = timezone.now()
        booking.clock_out_lat = lat
        booking.clock_out_lng = lng
        booking.clock_out_postcode = resolved_pc or None
        booking.clock_out_note = note
        booking.clock_out_supervisor_name = supervisor_name

        sig_file = _save_signature_from_dataurl(sig_data_url, filename_prefix=f"booking_{booking.id}_clockout")
        if sig_file:
            booking.clock_out_signature.save(sig_file.name, sig_file, save=False)

        booking.save()
        
        # For Audit log
        log_audit(actor=request.user, subject=request.user, action=AuditAction.CLOCK_OUT,
              shift=booking.shift, booking=booking,
              message="Clock out recorded.",
              detected_postcode=resolved_pc, note=note, supervisor=supervisor_name, lat=lat, lng=lng)

        messages.success(request, "Clock-out recorded.")
        return _clock_json(True, "Clock-out successful.")
        
    except Exception as e:
        # Log the error for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Clock-out error for booking {booking_id}: {str(e)}", exc_info=True)
        return _clock_json(False, "An error occurred during clock-out. Please try again.", status=500)

# ---------- Reports ----------
@login_required
@user_passes_test(lambda u: u.is_staff)
def attendance_report(request):
    # ---- org scope ----
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    # ---- date range ----
    today = timezone.localdate()
    default_start = today - timedelta(days=7)
    try:
        start_str = request.GET.get("start")
        end_str = request.GET.get("end")
        start = date.fromisoformat(start_str) if start_str else default_start
        end = date.fromisoformat(end_str) if end_str else today
    except ValueError:
        start, end = default_start, today

    include_paid = request.GET.get("include_paid") == "1"

    # ---- base queryset (org + date range) ----
    qs = (
        ShiftBooking._base_manager  # bypass any implicit tenant filtering
        .select_related("user", "shift")
        .filter(organization=tenant)
        .filter(shift__date__gte=start, shift__date__lte=end)
    )
    # apply paid filter only when requested
    if not include_paid:
        qs = qs.filter(paid_at__isnull=True)

    qs = qs.order_by("shift__date", "shift__start_time", "user__username")

    # ---- rows ----
    rows = []
    now = timezone.localtime()
    tz = timezone.get_current_timezone()
    for b in qs:
        sh = b.shift
        sd = datetime.combine(sh.date, sh.start_time or time.min, tzinfo=tz)
        ed = datetime.combine(sh.date, sh.end_time or time.max, tzinfo=tz)

        ci = b.clock_in_at
        co = b.clock_out_at
        ci_local = timezone.localtime(ci) if ci else None
        co_local = timezone.localtime(co) if co else None

        worked_sec = int((co - ci).total_seconds()) if ci and co and co >= ci else 0
        late_min   = int(round((ci - sd).total_seconds() / 60.0)) if ci and ci > sd else 0
        early_min  = int(round((ed - co).total_seconds() / 60.0)) if co and co < ed else 0

        if not ci and now > ed:
            status = "No show"
        elif ci and not co and now > ed:
            status = "Incomplete"
        elif ci and co:
            status = "Present"
        else:
            status = "Pending"

        hours = worked_sec // 3600
        minutes = (worked_sec % 3600) // 60
        worked_hhmm = f"{hours:02d}:{minutes:02d}"

        rows.append({
            "date": sh.date.isoformat(),
            "user": b.user.get_username(),
            "role": sh.role,
            "title": sh.title,
            "start": sd.strftime("%H:%M"),
            "end": ed.strftime("%H:%M"),
            "clock_in": ci_local.strftime("%H:%M") if ci_local else "",
            "clock_out": co_local.strftime("%H:%M") if co_local else "",
            "worked": worked_hhmm,
            "late_min": late_min,
            "early_min": early_min,
            "status": status,
        })

    fmt = (request.GET.get("format") or "").lower()
    if fmt == "csv":
        return _attendance_csv(rows, start, end)
    if fmt == "xlsx":
        try:
            return _attendance_xlsx(rows, start, end)
        except ImportError:
            messages.error(request, "Excel export requires 'openpyxl'. Try CSV export instead.")

    return render(request, "reports/attendance.html", {
        "rows": rows,
        "start": start,
        "end": end,
    })

def _attendance_csv(rows, start, end):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="attendance_{start}_{end}.csv"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "Date",
            "User",
            "Role",
            "Title",
            "Start",
            "End",
            "Clock In",
            "Clock Out",
            "Worked (HH:MM)",
            "Late (min)",
            "Early Leave (min)",
            "Status",
        ]
    )
    for r in rows:
        writer.writerow(
            [
                r["date"],
                r["user"],
                r["role"],
                r["title"],
                r["start"],
                r["end"],
                r["clock_in"],
                r["clock_out"],
                r["worked"],
                r["late_min"],
                r["early_min"],
                r["status"],
            ]
        )
    return response

def _attendance_xlsx(rows, start, end):
    # pip install openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = [
        "Date",
        "User",
        "Role",
        "Title",
        "Start",
        "End",
        "Clock In",
        "Clock Out",
        "Worked (HH:MM)",
        "Late (min)",
        "Early Leave (min)",
        "Status",
    ]
    ws.append(headers)

    for r in rows:
        ws.append(
            [
                r["date"],
                r["user"],
                r["role"],
                r["title"],
                r["start"],
                r["end"],
                r["clock_in"],
                r["clock_out"],
                r["worked"],
                r["late_min"],
                r["early_min"],
                r["status"],
            ]
        )

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="attendance_{start}_{end}.xlsx"'
    return response

# ---------- Admin dashboard ----------
@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    # ---- resolve active tenant/org (same pattern we used in admin_manage_shifts) ----
    tenant = getattr(request, "tenant", None)
    if tenant is None:
        tenant = getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    # ---- dates/times ----
    now = timezone.localtime()
    today = now.date()
    current_time = now.time()

    # upcoming = strictly in future OR later today (not finished)
    future_q = (
        Q(date__gt=today) |
        (Q(date=today) & (
            Q(end_time__gt=current_time) |
            Q(end_time__isnull=True, start_time__gt=current_time) |
            Q(start_time__isnull=True, end_time__isnull=True)
        ))
    )

    # ---- KPIs (all org-scoped) ----
    total_shifts = Shift._base_manager.filter(organization=tenant).count()
    total_bookings = ShiftBooking._base_manager.filter(organization=tenant).count()
    total_users = get_user_model().objects.filter(profile__organization=tenant).count()

    # available shifts = upcoming AND not full
    available_shifts = (
        Shift._base_manager
        .filter(future_q, organization=tenant)
        .annotate(booked_total=Count("bookings"))
        .filter(booked_total__lt=F("max_staff"))
        .count()
    )

    # today's shifts (any with date=today)
    todays_shifts = (
        Shift._base_manager
        .filter(organization=tenant, date=today)
        .count()
    )

    # ---- Upcoming table (org-scoped) ----
    upcoming_shifts = (
        Shift._base_manager
        .filter(future_q, organization=tenant)
        .annotate(booked_count=Count("bookings"))
        .order_by("date", "start_time")[:10]
    )

    # ---- Recent users list (org-scoped) ----
    User = get_user_model()
    users = (
        User.objects
        .filter(profile__organization=tenant)
        .order_by("-date_joined" if hasattr(User, "date_joined") else "-id")[:10]
    )

    context = {
        "available_shifts": available_shifts,
        "total_bookings": total_bookings,
        "total_users": total_users,
        "todays_shifts": todays_shifts,
        "upcoming_shifts": upcoming_shifts,
        "users": list(users),
        # (optional extras you had before)
        "total_shifts": total_shifts,
    }
    return render(request, "admin/dashboard.html", context)

# ---------- Admin: basic user management ----------
@user_passes_test(is_admin)
def admin_user_list(request):
    q = request.GET.get("q", "").strip()
    users = User.objects.all().order_by("-date_joined" if hasattr(User, "date_joined") else "id")
    if q:
        users = users.filter(
            Q(username__icontains=q)
            | Q(email__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
        )
    return render(request, "admin/users/list.html", {"users": users, "q": q})

@user_passes_test(is_admin)
def admin_user_create(request):
    if request.method == "POST":
        form = AdminUserCreateForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)  # type: ignore
            raw = form.cleaned_data.get("password") or _generate_temp_password()
            user.set_password(raw)
            user.save()
            messages.success(request, f"User '{user.username}' created.")
            return render(
                request,
                "admin/users/create_done.html",
                {"user_obj": user, "temp_password": raw},
            )
    else:
        form = AdminUserCreateForm()
    return render(request, "admin/users/create.html", {"form": form})

@login_required
@user_passes_test(is_admin)
def admin_user_send_reset(request, user_id):
    try:
        user = User.objects.get(pk=user_id)
    except (User.DoesNotExist, ValueError):
        messages.error(request, "User not found.")
        return redirect("admin_user_list")

    if not user.email:
        messages.error(request, "That user has no email address on file.")
        return redirect("admin_user_list")

    form = PasswordResetForm({"email": user.email})
    if form.is_valid():
        form.save(
            request=request,
            use_https=request.is_secure(),
            from_email=None,
        )
        messages.success(request, f"Password reset email sent to {user.email}.")
    else:
        messages.error(request, "Could not send reset email (invalid email).")
    return redirect("admin_user_list")

def _generate_temp_password(length=12):
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


@login_required
@user_passes_test(lambda u: u.is_staff)   # admins/staff only
def admin_manage_shifts(request):
    """
    Manage upcoming shifts and book users onto them, showing per-user compliance.
    """
    # ---- filters (mirror template fields) ----
    title_q = (request.GET.get("title_q") or "").strip()
    role_q  = (request.GET.get("role") or "").strip()
    start_q = request.GET.get("start") or ""
    end_q   = request.GET.get("end") or ""
    only_open = request.GET.get("only_open") == "1"
    user_q = (request.GET.get("user_q") or "").strip()

    # ---- resolve active tenant/org ----
    tenant = getattr(request, "tenant", None)
    if tenant is None:
        tenant = getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    # ---- base: upcoming (today onwards, not finished) ----
    now = timezone.localtime()
    today = now.date()
    current_time = now.time()
    future_q = (
        Q(date__gt=today) |
        (Q(date=today) & (Q(end_time__gt=current_time) | Q(end_time__isnull=True) | Q(start_time__gt=current_time)))
    )

    # shifts list (org-scoped)
    shifts_qs = (
        Shift._base_manager  # bypass TenantManager to avoid any hidden filters
        .filter(future_q, organization=tenant)
        .annotate(booked_total=Count("bookings"))
        .order_by("date", "start_time", "title")
    )

    if title_q:
        shifts_qs = shifts_qs.filter(title__icontains=title_q)
    if role_q:
        shifts_qs = shifts_qs.filter(role=role_q)
    if start_q:
        shifts_qs = shifts_qs.filter(date__gte=start_q)
    if end_q:
        shifts_qs = shifts_qs.filter(date__lte=end_q)
    if only_open:
        shifts_qs = shifts_qs.filter(booked_total__lt=F("max_staff"))

    upcoming_shifts = list(shifts_qs[:50])

    # users dropdown (org-scoped) — adjust if staff can book cross-org
    users_qs = (
        User.objects
        .filter(is_active=True, profile__organization=tenant)
        .order_by("username")
    )
    if user_q:
        users_qs = users_qs.filter(
            Q(username__icontains=user_q) |
            Q(email__icontains=user_q) |
            Q(first_name__icontains=user_q) |
            Q(last_name__icontains=user_q)
        )
    users = list(users_qs[:300])

    # attach (user, is_compliant) rows
    for sh in upcoming_shifts:
        sh.user_rows = [(u, user_is_compliant_for_role(u, sh.role)) for u in users]

    # ---- metrics (ALL org-scoped so they match the list) ----
    total_shifts = Shift._base_manager.filter(organization=tenant).count()
    total_users  = User.objects.filter(profile__organization=tenant).count()
    available_upcoming = (
        Shift._base_manager.filter(future_q, organization=tenant)
        .annotate(c=Count("bookings"))
        .filter(c__lt=F("max_staff"))
        .count()
    )
    booked_upcoming = (
        Shift._base_manager.filter(future_q, organization=tenant)
        .annotate(c=Count("bookings"))
        .filter(c__gt=0)
        .count()
    )

    # recent bookings (org-scoped)
    recent_bookings = (
        ShiftBooking._base_manager
        .select_related("user", "shift")
        .filter(organization=tenant)
        .order_by("-id")[:50]
    )

    context = {
        "total_shifts": total_shifts,
        "total_users": total_users,
        "available_upcoming": available_upcoming,
        "booked_upcoming": booked_upcoming,
        "upcoming_shifts": upcoming_shifts,
        "users": users,
        "title_q": title_q,
        "role_q": role_q,
        "start_q": start_q,
        "end_q": end_q,
        "only_open": only_open,
        "user_q": user_q,
        "recent_bookings": recent_bookings,
    }
    return render(request, "admin/manage_shifts.html", context)


# ---------- ADMIN: Paid bookings ----------
@login_required
@user_passes_test(is_admin)
def admin_paid_bookings(request):
    """
    List all PAID bookings with filters + CSV/XLSX export (same shape as attendance).
    """
    # resolve active org/tenant
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    today = timezone.localdate()
    default_start = today - timedelta(days=30)

    # filters
    start_str = request.GET.get("start") or ""
    end_str = request.GET.get("end") or ""
    user_q = (request.GET.get("user_q") or "").strip()
    role_q = (request.GET.get("role") or "").strip()

    try:
        start = date.fromisoformat(start_str) if start_str else default_start
        end = date.fromisoformat(end_str) if end_str else today
    except ValueError:
        start, end = default_start, today

    qs = (
        ShiftBooking._base_manager                # bypass any tenant manager
        .select_related("user", "shift")
        .filter(organization=tenant)             # scope to active org
        .filter(paid_at__isnull=False)
        .filter(shift__date__gte=start, shift__date__lte=end)
        .order_by("-paid_at", "-id")
    )
    if user_q:
        qs = qs.filter(
            Q(user__username__icontains=user_q) |
            Q(user__first_name__icontains=user_q) |
            Q(user__last_name__icontains=user_q) |
            Q(user__email__icontains=user_q)
        )
    if role_q:
        qs = qs.filter(shift__role__iexact=role_q)

    # build rows (reuse attendance layout + paid_at column)
    rows = []
    for b in qs:
        sh = b.shift
        sd = datetime.combine(sh.date, sh.start_time or time.min, tzinfo=timezone.get_current_timezone())
        ed = datetime.combine(sh.date, sh.end_time or time.max, tzinfo=timezone.get_current_timezone())
        worked_sec = int((b.clock_out_at - b.clock_in_at).total_seconds()) if b.clock_in_at and b.clock_out_at and b.clock_out_at >= b.clock_in_at else 0
        hours = worked_sec // 3600
        minutes = (worked_sec % 3600) // 60
        rows.append({
            "date": sh.date.isoformat(),
            "user": b.user.get_username(),
            "role": sh.role,
            "title": sh.title,
            "start": sd.strftime("%H:%M"),
            "end": ed.strftime("%H:%M"),
            "clock_in": timezone.localtime(b.clock_in_at).strftime("%H:%M") if b.clock_in_at else "",
            "clock_out": timezone.localtime(b.clock_out_at).strftime("%H:%M") if b.clock_out_at else "",
            "worked": f"{hours:02d}:{minutes:02d}",
            "paid_at": timezone.localtime(b.paid_at).strftime("%Y-%m-%d %H:%M"),
            "booking_id": b.id,   # <— add this
            
        })

    fmt = (request.GET.get("format") or "").lower()
    if fmt == "csv":
        # include paid_at in export
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="paid_{start}_{end}.csv"'
        import csv
        writer = csv.writer(response)
        writer.writerow(["Date","User","Role","Title","Start","End","Clock In","Clock Out","Worked (HH:MM)","Paid At"])
        for r in rows:
            writer.writerow([r["date"],r["user"],r["role"],r["title"],r["start"],r["end"],r["clock_in"],r["clock_out"],r["worked"],r["paid_at"]])
        return response

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            ws = wb.active
            ws.title = "Paid"
            headers = ["Date","User","Role","Title","Start","End","Clock In","Clock Out","Worked (HH:MM)","Paid At"]
            ws.append(headers)
            for r in rows:
                ws.append([r["date"],r["user"],r["role"],r["title"],r["start"],r["end"],r["clock_in"],r["clock_out"],r["worked"],r["paid_at"]])
            for col in range(1, len(headers)+1):
                ws.column_dimensions[get_column_letter(col)].width = 18
            from io import BytesIO
            bio = BytesIO(); wb.save(bio); bio.seek(0)
            response = HttpResponse(bio.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = f'attachment; filename="paid_{start}_{end}.xlsx"'
            return response
        except ImportError:
            messages.error(request, "Excel export requires 'openpyxl'. Try CSV instead.")

    return render(request, "admin/paid_bookings.html", {
        "rows": rows,
        "start": start,
        "end": end,
        "user_q": user_q,
        "role_q": role_q,
        "count": qs.count(),
    })


# ---------- ADMIN: mark / unmark paid ----------
@require_POST
@login_required
@user_passes_test(is_admin)
def admin_mark_paid(request, booking_id):
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    b = get_object_or_404(ShiftBooking._base_manager, pk=booking_id, organization=tenant)
    if not b.paid_at:
        b.paid_at = timezone.now()
        b.save(update_fields=["paid_at"])
        messages.success(request, f"Marked booking #{b.id} as paid.")
    return redirect(request.META.get("HTTP_REFERER") or "admin_dashboard")

@require_POST
@login_required
@user_passes_test(is_admin)
def admin_unmark_paid(request, booking_id):
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    b = get_object_or_404(ShiftBooking._base_manager, pk=booking_id, organization=tenant)
    if b.paid_at:
        b.paid_at = None
        b.save(update_fields=["paid_at"])
        messages.success(request, f"Unmarked booking #{b.id} as paid.")
    return redirect(request.META.get("HTTP_REFERER") or "admin_dashboard")


# ---------- USER: My paid shifts ----------
@login_required
def my_paid_shifts(request):
    # resolve tenant the same way as my_bookings
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    bookings = (
        ShiftBooking._base_manager
        .select_related("shift")
        .filter(user=request.user, organization=tenant, paid_at__isnull=False)
        .order_by("-paid_at", "-id")
    )
    return render(request, "bookings/paid.html", {"bookings": bookings})


@require_POST
@login_required
@user_passes_test(lambda u: u.is_staff)
def admin_book_for_user(request):
    """
    Books a selected user onto a shift from the Manage Shifts page.
    Enforces compliance. If you want an override checkbox, look for 'override' in POST.
    """
    try:
        shift_id = int(request.POST.get("shift_id") or "0")
        user_id  = int(request.POST.get("user_id") or "0")
    except ValueError:
        messages.error(request, "Invalid parameters.")
        return redirect("admin_manage_shifts")

    shift = get_object_or_404(Shift, id=shift_id)
    user  = get_object_or_404(User, id=user_id)

    # Optional override (unchecked by default)
    override = request.POST.get("override") == "1"
    reason   = (request.POST.get("reason") or "").strip()

    # Capacity check
    booked_total = ShiftBooking.objects.filter(shift=shift).count()
    if booked_total >= shift.max_staff:
        messages.error(request, f"'{shift.title}' is already full.")
        return redirect("admin_manage_shifts")

    # Duplicate check
    if ShiftBooking.objects.filter(shift=shift, user=user).exists():
        messages.info(request, f"{user} is already booked on '{shift.title}'.")
        return redirect("admin_manage_shifts")

    # Compliance guard
    if not user_is_compliant_for_role(user, shift.role) and not override:
        messages.error(
            request,
            f"{user.get_username()} is not compliant for role '{shift.get_role_display()}'. "
            "Tick 'Override' and provide a reason to proceed."
        )
        return redirect("admin_manage_shifts")

    # In admin_book_for_user, ensure organization is set
    from django.db import IntegrityError
    try:
        booking = ShiftBooking.objects.create(user=user, shift=shift, organization=shift.organization)
    except IntegrityError:
        messages.info(request, f"{user} is already booked on '{shift.title}'.")
        return redirect("admin_manage_shifts")
    note = f" (override: {reason})" if override and reason else (" (override)" if override else "")
    messages.success(request, f"Booked {user.get_username()} on '{shift.title}'.{note}")
    
    # For Audit log
    log_audit(actor=request.user, subject=user, action=AuditAction.BOOKING_CREATED,
          shift=shift, booking=booking,
          message=f"Admin assigned {user.username} to '{shift.title}'.",
          admin_assign=True)
    
    return redirect("admin_manage_shifts")


@login_required
@user_passes_test(is_admin)
@require_POST
def admin_cancel_booking_admin(request, booking_id):
    """
    POST to cancel a booking by id.
    """
    org = _active_tenant(request)
    booking = get_object_or_404(ShiftBooking._base_manager, pk=booking_id, organization=org)
    title = booking.shift.title
    username = booking.user.get_username()
    booking.delete()
    
    # For Audit log
    log_audit(actor=request.user, subject=booking.user, action=AuditAction.BOOKING_CANCELLED,
          shift=booking.shift, booking=None,
          message=f"Admin assigned {request.user.username} cancelled booking for '{booking.shift.title}'.")

    messages.success(request, f"Cancelled booking for {username} on '{title}'.")
    return redirect("admin_manage_shifts")


@login_required
@user_passes_test(is_admin)
@require_POST
def admin_clock_in_for_user(request, booking_id):
    override = request.POST.get("override") == "1"
    reason = (request.POST.get("reason") or "").strip()

    org = _active_tenant(request)

    qs = ShiftBooking._base_manager.select_related("shift", "user")
    if org:
        qs = qs.filter(organization=org)   # only scope when known

    booking = get_object_or_404(qs, pk=booking_id)

    # (Optional guard: block cross-org unless superuser)
    if org and booking.organization_id != getattr(org, "id", None) and not request.user.is_superuser:
        messages.error(request, "That booking is not in the active workspace.")
        return redirect("admin_manage_shifts")

    if booking.clock_in_at:
        messages.info(request, "Already clocked in.")
        return redirect("admin_manage_shifts")

    booking.clock_in_at = timezone.now()
    booking.clock_out_note = (booking.clock_out_note or "")
    if reason:
        booking.clock_out_note += (("\n" if booking.clock_out_note else "") + f"[Admin IN] {reason}")
    booking.save(update_fields=["clock_in_at", "clock_out_note"])

    # ✅ Log the correct action
    log_audit(
        actor=request.user,
        subject=booking.user,
        action=AuditAction.CLOCK_IN,
        shift=booking.shift,
        booking=booking,
        message="Admin clock-in recorded.",
    )

    messages.success(request, "Clock-in recorded.")
    return redirect(request.META.get("HTTP_REFERER") or "admin_manage_shifts")


@login_required
@user_passes_test(is_admin)
@require_POST
def admin_clock_out_for_user(request, booking_id):
    override = request.POST.get("override") == "1"
    reason = (request.POST.get("reason") or "").strip()

    org = _active_tenant(request)

    qs = ShiftBooking._base_manager.select_related("shift", "user")
    if org:
        qs = qs.filter(organization=org)

    booking = get_object_or_404(qs, pk=booking_id)

    if org and booking.organization_id != getattr(org, "id", None) and not request.user.is_superuser:
        messages.error(request, "That booking is not in the active workspace.")
        return redirect("admin_manage_shifts")

    if booking.clock_out_at:
        messages.info(request, "Already clocked out.")
        return redirect("admin_manage_shifts")

    booking.clock_out_at = timezone.now()
    booking.clock_out_note = (booking.clock_out_note or "")
    if reason:
        booking.clock_out_note += (("\n" if booking.clock_out_note else "") + f"[Admin OUT] {reason}")
    booking.save(update_fields=["clock_out_at", "clock_out_note"])

    log_audit(
        actor=request.user,
        subject=booking.user,
        action=AuditAction.CLOCK_OUT,
        shift=booking.shift,
        booking=booking,
        message="Admin clock-out recorded.",
    )

    messages.success(request, "Clock-out recorded.")
    return redirect(request.META.get("HTTP_REFERER") or "admin_manage_shifts")

@require_POST
@login_required
@user_passes_test(is_admin)
def admin_mark_booking_paid(request, booking_id):
    org = _active_tenant(request)
    # Use _base_manager to bypass any implicit tenant filtering,
    # then explicitly scope to the active org.
    booking = get_object_or_404(
        ShiftBooking._base_manager,
        id=booking_id,
        organization=org,
    )

    if not booking.clock_in_at or not booking.clock_out_at:
        messages.error(request, "Cannot mark as paid until the shift is completed.")
    else:
        booking.mark_paid()  # sets paid_at = now
        messages.success(request, f"Booking #{booking.id} marked as paid.")

    return redirect(request.POST.get("next") or "admin_manage_shifts")

# Compliance / audit trail views could go here...

def is_admin(user):
    return user.is_authenticated and user.is_staff

# ---------- Admin: upload & list ----------
@login_required
@user_passes_test(is_admin)
def compliance_admin_upload(request):
    """
    Single-page admin uploader. Shows recent docs and a form to add a new one.
    """
    form = AdminComplianceUploadForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        doc = form.save(commit=False)
        doc.uploaded_by = request.user
        doc.save()
        messages.success(request, "Compliance document uploaded.")
        return redirect("compliance_admin_upload")

    # Optional quick filters
    q = request.GET.get("q", "").strip()
    user_id = request.GET.get("user") or ""
    dtype = request.GET.get("type") or ""
    status = request.GET.get("status") or ""

    docs = ComplianceDocument.objects.select_related("user", "doc_type").all()
    if q:
        docs = docs.filter(
            Q(user__username__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(doc_type__name__icontains=q)
        )
    if user_id.isdigit():
        docs = docs.filter(user_id=int(user_id))
    if dtype.isdigit():
        docs = docs.filter(doc_type_id=int(dtype))
    if status in {"pending", "approved", "rejected"}:
        docs = docs.filter(status=status)

    docs = docs.order_by("-uploaded_at")[:100]

    context = {
        "form": form,
        "docs": docs,
        "users": User.objects.order_by("username")[:200],
        "types": ComplianceDocType.objects.filter(is_active=True).order_by("name"),
        "q": q, "user_filter": user_id, "type_filter": dtype, "status_filter": status,
    }
    return render(request, "compliance/admin_upload.html", context)


# ---------- User: read-only ----------
@login_required
def my_compliance(request):
    docs = (
        ComplianceDocument.objects
        .select_related("doc_type")
        .filter(user=request.user)
        .order_by("doc_type__name", "-uploaded_at")
    )
    return render(request, "compliance/my_compliance.html", {"docs": docs})

# ---------- User Availability Views ----------
@login_required
def my_availability(request):
    """User can view and manage their availability"""
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    availability = (
        UserAvailability._base_manager
        .filter(user=request.user, organization=tenant)
        .order_by('date', 'start_time')
    )

    return render(request, "availability/my_availability.html", {"availability": availability})


@login_required
def add_availability(request):
    """Add new availability entry"""
    from .forms import UserAvailabilityForm
    from .models import UserAvailability
    
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    if request.method == 'POST':
        form = UserAvailabilityForm(request.POST)
        if form.is_valid():
            availability = form.save(commit=False)
            availability.user = request.user
            availability.organization = tenant
            try:
                availability.save()
                messages.success(request, "Availability added successfully.")
                return redirect("my_availability")
            except IntegrityError:
                messages.error(request, "You already have availability set for this date and time.")
        else:
            messages.error(request, "Please fix the errors below.")
    else:
        initial_date = request.GET.get('date')
        initial = {'date': initial_date} if initial_date else {}
        form = UserAvailabilityForm(initial=initial)

    return render(request, "availability/add_availability.html", {"form": form})


# ---------- Holiday Request Views ----------
@login_required
def my_holidays(request):
    """User can view their holiday requests"""
    from .models import HolidayRequest
    
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    holidays = (
        HolidayRequest._base_manager
        .filter(user=request.user, organization=tenant)
        .order_by('-created_at')
    )

    return render(request, "holidays/my_holidays.html", {"holidays": holidays})


@login_required
def request_holiday(request):
    """Create new holiday request"""
    from .forms import HolidayRequestForm
    from .models import HolidayRequest
    
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    if request.method == 'POST':
        form = HolidayRequestForm(request.POST)
        if form.is_valid():
            holiday = form.save(commit=False)
            holiday.user = request.user
            holiday.organization = tenant
            holiday.save()
            
            log_audit(
                actor=request.user,
                subject=request.user,
                action=AuditAction.BOOKING_CREATED,
                message=f"Holiday request submitted: {holiday.get_holiday_type_display()} from {holiday.start_date} to {holiday.end_date}"
            )
            
            messages.success(request, "Holiday request submitted successfully.")
            return redirect("my_holidays")
        else:
            messages.error(request, "Please fix the errors below.")
    else:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        initial = {}
        if start_date:
            initial['start_date'] = start_date
        if end_date:
            initial['end_date'] = end_date
        form = HolidayRequestForm(initial=initial)

    return render(request, "holidays/request_holiday.html", {"form": form})


# ---------- Admin Holiday Management Views ----------
@login_required
@user_passes_test(is_admin)
def admin_holiday_requests(request):
    """Admin view to manage all holiday requests"""
    from .models import HolidayRequest
    
    tenant = getattr(request, "tenant", None) or getattr(getattr(request.user, "profile", None), "organization", None)
    if tenant is None:
        messages.error(request, "No active workspace selected. Please select an organization.")
        return redirect("home")

    status_filter = request.GET.get('status', '')
    user_filter = request.GET.get('user', '')
    
    holidays = HolidayRequest._base_manager.select_related('user').filter(organization=tenant)
    
    if status_filter:
        holidays = holidays.filter(status=status_filter)
    if user_filter:
        holidays = holidays.filter(user__username__icontains=user_filter)
    
    holidays = holidays.order_by('-created_at')

    context = {
        'holidays': holidays,
        'status_filter': status_filter,
        'user_filter': user_filter,
        'status_choices': HolidayRequest.STATUS_CHOICES,
    }
    return render(request, "admin/holiday_requests.html", context)